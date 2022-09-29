# Author: CDamogu
# Date: 2022/05/12
# Des: Set Border to Excel Singel Cells, openpyxl

import openpyxl
from openpyxl.styles import *

outputXlName = '测试.xlsx'
inputShtName = 'Sheet_test'

wb = openpyxl.load_workbook(outputXlName)
ws = wb[inputShtName]


# style = NoneSet(values=('dashDot','dashDotDot', 'dashed','dotted',
#                             'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
#                             'mediumDashed', 'slantDashDot', 'thick', 'thin')
#                     )
# 设置边框{'medium' 中粗 'thin'  细  'thick'  粗  'dashed'  虚线  'dotted'  点线}
def format_border_cell(ws, row_index, col_index):
    ws.cell(row_index, col_index).border = Border(top=Side(border_style='hair', color='FF000000'),
                                                  right=Side(border_style='hair', color='FF000000'),
                                                  bottom=Side(border_style='hair', color='FF000000'),
                                                  left=Side(border_style='hair', color='FF000000'))
    font_set = Font(name='等线', size=12, color=colors.BLACK, bold=True)
    ws.cell(row_index, col_index).font = font_set
    ws.cell(row_index, col_index).alignment = Alignment(horizontal='center', vertical='center')


if __name__ == '__main__':
    # 调用函数，给('A1')单元格设置边框
    format_border_cell(ws, 2, 2)

    wb.save(outputXlName)
