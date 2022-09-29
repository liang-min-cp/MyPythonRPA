def pyxl_execl():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '麦嘎'
    data = [(1, 2, 3, 4), ('djdj', 'dkdk', '', 'djdj'), (23, 98, 'dk', 'kdkd')]
    fill = openpyxl.styles.PatternFill("solid", fgColor="CCECFF")  # 颜色也可以直接设置red等

    row0 = ('姓名', 'name', 'age', 'core')
    for i, r in enumerate(row0):
        print(i,r)
        ws.cell(row=1, column=i + 1, value=r)
    print("========================================")
    for i, r in enumerate(data):
        print(i,r)
        for j, c in enumerate(r):
            print(j,c)
            if i % 2 == 1:
                ws.cell(row=i + 2, column=j + 1).fill = fill

            ws.cell(row=i + 2, column=j + 1, value=c)

    wb.save('./测试.xlsx')


if __name__ == '__main__':

    pyxl_execl()