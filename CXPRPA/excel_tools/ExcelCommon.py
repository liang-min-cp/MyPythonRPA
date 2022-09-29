import os
import sys
import openpyxl
from openpyxl.cell.cell import *
from openpyxl.styles import *

sys.path.append("..")


def GetExcelTemplateFullPath(file_name: str):
    cur_path = os.path.abspath(os.path.dirname(__file__))
    file_dictionary = cur_path.split("CXPRPA")[0] + "\\CXPRPA\\excel_template"
    path = file_dictionary + "\\" + file_name

    full_path_old = sys.argv[0]
    full_path_old_list = full_path_old.split(R"/")
    list_last_element = R"excel_template/" + file_name
    full_path_old_list[full_path_old_list.__len__() - 1] = list_last_element
    path_new = R"/".join(full_path_old_list)
    return path_new


def GetExcelResultFullPath():
    cur_path = os.path.abspath(os.path.dirname(__file__))
    file_dictionary = cur_path.split("CXPRPA")[0] + "\\CXPRPA\\result"
    # 用日期流水号作为输出文件名
    path = file_dictionary + "\\" + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":", "-") + ".xlsx"

    full_path_old = sys.argv[0]
    full_path_old_list = full_path_old.split(R"/")
    list_last_element = R"result/" + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S").replace(":", "-") + ".xlsx"
    full_path_old_list[full_path_old_list.__len__() - 1] = list_last_element
    path_new = R"/".join(full_path_old_list)

    return path_new


def GetExcelConfigFullPath(dic_name: str, file_name: str):
    full_path_old = sys.argv[0]
    full_path_old_list = full_path_old.split(R"/")
    list_last_element = dic_name + "/" + file_name
    full_path_old_list[full_path_old_list.__len__() - 1] = list_last_element
    path_new = R"/".join(full_path_old_list)
    return path_new


# 设置表头单元格样式
# style = NoneSet(values=('dashDot','dashDotDot', 'dashed','dotted',
#                             'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
#                             'mediumDashed', 'slantDashDot', 'thick', 'thin')
#                     )
# 设置边框{'medium' 中粗 'thin'  细  'thick'  粗  'dashed'  虚线  'dotted'  点线}
def SetHeaderCellStyle(cell: Cell):
    cell.border = Border(top=Side(border_style='hair', color='FF000000'),
                         right=Side(border_style='hair', color='FF000000'),
                         bottom=Side(border_style='hair', color='FF000000'),
                         left=Side(border_style='hair', color='FF000000'))
    cell.font = Font(name='等线', size=12, color=colors.BLACK, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')


# 设置表体单元格样式
def SetBodyCellStyle(cell: Cell):
    cell.border = Border(top=Side(border_style='hair', color='FF000000'),
                         right=Side(border_style='hair', color='FF000000'),
                         bottom=Side(border_style='hair', color='FF000000'),
                         left=Side(border_style='hair', color='FF000000'))
    cell.font = Font(name='等线', size=11, color=colors.BLACK, bold=False)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="CCECFF")  # 颜色


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'test'
    ws.cell(2, 2).value = "实际值2"
    ws.cell(3, 3).value = "OK"
    ws.cell(4, 4).value = "5446488646"
    SetHeaderCellStyle(ws.cell(2, 2))
    SetBodyCellStyle(ws.cell(3, 3))
    SetBodyCellStyle(ws.cell(4, 4))
    wb.save('./测试.xlsx')
