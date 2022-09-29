import datetime
import os
import sys

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_tools.ExcelCommon import SetHeaderCellStyle, GetExcelTemplateFullPath, GetExcelResultFullPath, \
    SetBodyCellStyle
from excel_tools.ExcelRead import ExcelRead

sys.path.append("..")
from log.logger import logger
from models.ProcessAndResult import ProcessAndResult


# 初始化明细表头信息
def InitTableHeader(excel_wb: Worksheet, process_and_result: ProcessAndResult):
    wb = excel_wb
    add_col_nums = process_and_result.Test_Circulate_Number  # 测量循环次数， 添加的表头列数
    real_val_header_text_model = "实际值"
    test_result_text_model = "测试结果"
    start_col_num = process_and_result.Coordinate.Start_Col_Num  # 起始列数 结果测量,此处默认9
    detail_table_head_row_num = process_and_result.Coordinate.Detail_Table_Head_Row_Num  # 表头行号
    for i in range(add_col_nums):
        real_val_col_num = start_col_num + i * 2
        test_result_col_num = start_col_num + i * 2 + 1
        real_val_text = real_val_header_text_model + (i + 1).__str__()
        test_result_text = test_result_text_model + (i + 1).__str__()
        real_val_cell = wb.cell(detail_table_head_row_num, real_val_col_num)
        test_result_cell = wb.cell(detail_table_head_row_num, test_result_col_num)
        real_val_cell.value = real_val_text
        test_result_cell.value = test_result_text
        SetHeaderCellStyle(real_val_cell)
        SetHeaderCellStyle(test_result_cell)


# 填写测试结果数据进入excel
def WriteSummaryToExcel(excel_wb: Worksheet, process_and_result: ProcessAndResult):
    wb = excel_wb
    # 填写的数据
    test_command_all_num = process_and_result.TestCommandAllNum  # 测试命令总条数
    test_command_success_num = process_and_result.TestCommandSuccessNum  # 测试命令成功条数
    test_command_failed_num = process_and_result.TestCommandFailedNum  # 测试命令失败条数
    test_time_all = process_and_result.TestTimeALl  # 测试命令总耗时
    test_time_average = process_and_result.TestTimeAverage  # 测试命令平均耗时
    test_result_summary = process_and_result.TestResultSummary  # 汇总本次测量总结果

    # 坐标信息
    test_command_all_num_Location = process_and_result.Coordinate.TestCommandAllNum_Location  # 测试命令总条数, 坐标
    test_command_success_num_Location = process_and_result.Coordinate.TestCommandSuccessNum_Location  # 测试命令成功条数, 坐标
    test_command_failed_num_Location = process_and_result.Coordinate.TestCommandFailedNum_Location  # 测试命令失败条数, 坐标
    test_time_all_Location = process_and_result.Coordinate.TestTimeALl_Location  # 测试命令总耗时, 坐标
    test_time_average_Location = process_and_result.Coordinate.TestTimeAverage_Location  # 测试命令平均耗时, 坐标
    test_result_summary_Location = process_and_result.Coordinate.TestResultSummary_Location  # 汇总本次测量总结果

    # 写入excel
    test_command_all_num_cell = wb.cell(test_command_all_num_Location[0], test_command_all_num_Location[1])
    test_command_success_num_cell = wb.cell(test_command_success_num_Location[0], test_command_success_num_Location[1])
    test_command_failed_num_cell = wb.cell(test_command_failed_num_Location[0], test_command_failed_num_Location[1])
    test_time_all_cell = wb.cell(test_time_all_Location[0], test_time_all_Location[1])
    test_time_average_cell = wb.cell(test_time_average_Location[0], test_time_average_Location[1])
    test_result_summary_cell = wb.cell(test_result_summary_Location[0], test_result_summary_Location[1])

    test_command_all_num_cell.value = test_command_all_num
    test_command_success_num_cell.value = test_command_success_num
    test_command_failed_num_cell.value = test_command_failed_num
    test_time_all_cell.value = test_time_all
    test_time_average_cell.value = test_time_average
    test_result_summary_cell.value = test_result_summary

    SetBodyCellStyle(test_command_all_num_cell)
    SetBodyCellStyle(test_command_success_num_cell)
    SetBodyCellStyle(test_command_failed_num_cell)
    SetBodyCellStyle(test_time_all_cell)
    SetBodyCellStyle(test_time_average_cell)
    SetBodyCellStyle(test_result_summary_cell)

    logger.info("WriteSummaryToExcel success")


# 填写测试过程数据进入excel
def WriteProcessToExcel(excel_wb: Worksheet, process_and_result: ProcessAndResult):
    wb = excel_wb
    for Command in process_and_result.Commands:
        logger.info(Command.__dict__)
        # 写入测试结果
        test_result = Command.Test_Result  # 测试结果
        test_result_Location = Command.Coordinate.Test_Result_Location  # 测试结果 坐标
        test_result_cell = wb.cell(test_result_Location[0], test_result_Location[1])
        test_result_cell.value = test_result
        SetBodyCellStyle(test_result_cell)

        # 写入测试明细, 多次测量每一次测量结果
        for Test_Res in Command.Test_Results:
            # 坐标
            test_actual_value_Location = Test_Res.Coordinate.Test_Actual_Value_Location  # 实际值 坐标
            test_actual_result_Location = Test_Res.Coordinate.Test_Actual_Result_Location  # 测试结果 坐标
            test_actual_value = Test_Res.Test_Actual_Value  # 实际值
            test_actual_result = Test_Res.Test_Actual_Result  # 测试结果
            test_actual_value_cell = wb.cell(test_actual_value_Location[0], test_actual_value_Location[1])
            test_actual_result_cell = wb.cell(test_actual_result_Location[0], test_actual_result_Location[1])
            test_actual_value_cell.value = test_actual_value
            test_actual_result_cell.value = test_actual_result
            SetBodyCellStyle(test_actual_value_cell)
            SetBodyCellStyle(test_actual_result_cell)

    logger.info("WriteProcessToExcel success")


class ExcelWrite:

    def __init__(self, workbook: Workbook, process_and_result: ProcessAndResult):
        self.excel = workbook
        self.Process_And_Result = process_and_result
        logger.info("Init")

    # 初始化明细表头

    # 把测试结果写入Excel
    def WriteProcessAndResultInfoToExcel(self):
        sheets = self.excel.sheetnames
        current_sheet_name = self.Process_And_Result.Coordinate.Sheet_Name
        if sheets.__contains__(current_sheet_name):
            sheet = self.excel[current_sheet_name]
            logger.info(sheet)
            # (1) 初始化明细表头信息，添加实际值1，测试结果2，实际值2...
            InitTableHeader(sheet, self.Process_And_Result)
            # (2) 填写数据进入excel
            # 2.1 写测试结果数据
            WriteSummaryToExcel(sheet, self.Process_And_Result)
            # 2.2 写测试过程数据
            WriteProcessToExcel(sheet, self.Process_And_Result)
        else:
            logger.info("can not find sheet : " + current_sheet_name)
            return None


if __name__ == '__main__':
    # 打开Excel
    excel_template_file_name = "TCsetting_Template.xlsx"
    file_path = GetExcelTemplateFullPath(excel_template_file_name)
    logger.info(file_path)
    excel_workbook = openpyxl.load_workbook(file_path)  # 获取表格文件

    # 读取Excel
    excelRead = ExcelRead(excel_workbook)
    processAndResul = excelRead.ReadProcessAndResultInfo()
    logger.info(processAndResul.__dict__)

    # 写入Excel
    excelWrite = ExcelWrite(excel_workbook, processAndResul)
    excelWrite.WriteProcessAndResultInfoToExcel()

    # 关闭Excel
    file_path_output = GetExcelResultFullPath()
    if os.path.isfile(file_path_output):
        os.remove(file_path_output)
    logger.info(file_path_output)
    excel_workbook.save(file_path_output)
    excel_workbook.close()
