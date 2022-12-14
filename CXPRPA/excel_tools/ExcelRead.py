import sys

import openpyxl
from openpyxl import Workbook

sys.path.append("..")
from log.logger import logger
from excel_tools.SystemConfigCoordinate import SystemConfigCoordinate
from models.SysConfig import SysConfigModel
from excel_tools.SummaryCoordinate import SummaryCoordinate
from models.ProcessResult import ProcessResult
from models.ProcessCommand import ProcessCommand
from models.ProcessAndResult import ProcessAndResult
from excel_tools.CommandCoordinate import CommandCoordinate
from excel_tools.ResultCoordinate import ResultCoordinate
from excel_tools.ExcelCommon import GetExcelTemplateFullPath


class ExcelRead:

    def __init__(self, workbook: Workbook):
        self.excel = workbook

    def ReadSystemConfigInfo(self):
        coordinate = SystemConfigCoordinate()
        logger.info(coordinate.__dict__)
        sheets = self.excel.sheetnames
        if sheets.__contains__(coordinate.SheetName):
            sheet = self.excel[coordinate.SheetName]
            logger.info(sheet)
            fins_trans_way = sheet.cell(coordinate.Fins_Trans_Way_Location[0],
                                        coordinate.Fins_Trans_Way_Location[1]).value
            ip_computer = sheet.cell(coordinate.Ip_Computer_Location[0], coordinate.Ip_Computer_Location[1]).value
            ip_plc = sheet.cell(coordinate.Ip_Plc_Location[0], coordinate.Ip_Plc_Location[1]).value
            port = sheet.cell(coordinate.Port_Location[0], coordinate.Port_Location[1]).value
            logger.info(fins_trans_way)
            logger.info(ip_computer)
            logger.info(ip_plc)
            logger.info(port)
            sys_config_model = SysConfigModel(fins_trans_way, ip_computer, ip_plc, port, coordinate)
            # logger.info(sys_config_model.__dict__)
            return sys_config_model
        else:
            logger.info("can not find sheet : " + coordinate.SheetName)
            return None

    def ReadProcessAndResultInfo(self):
        coordinate = SummaryCoordinate()
        sheets = self.excel.sheetnames
        if sheets.__contains__(coordinate.Sheet_Name):
            sheet = self.excel[coordinate.Sheet_Name]
            logger.info(sheet)
            # ??????????????????
            test_circulate_number = sheet.cell(coordinate.Test_Circulate_Number_Location[0], coordinate.Test_Circulate_Number_Location[1]).value
            # CXP????????????
            cxp_file_path = sheet.cell(coordinate.Cxp_File_Path_Location[0], coordinate.Cxp_File_Path_Location[1]).value
            # CX-Programmer?????????
            cx_programmer_version = sheet.cell(coordinate.CX_Programmer_Version_Location[0], coordinate.CX_Programmer_Version_Location[1]).value
            # RX65???????????????
            rx65_version = sheet.cell(coordinate.RX65_Version_Location[0], coordinate.RX65_Version_Location[1]).value
            # AM3352???????????????
            am3352_version = sheet.cell(coordinate.AM3352_Version_Location[0], coordinate.AM3352_Version_Location[1]).value

            logger.info("?????????????????? : " + str(test_circulate_number))
            logger.info("CXP???????????? : " + str(cxp_file_path))
            logger.info("CX-Programmer????????? : " + str(cx_programmer_version))
            logger.info("RX65??????????????? : " + str(rx65_version))
            logger.info("AM3352??????????????? : " + str(am3352_version))
            processAndResult = ProcessAndResult(test_circulate_number, cxp_file_path, coordinate)
            processAndResult.CX_Programmer_Version = cx_programmer_version
            processAndResult.RX65_Version = rx65_version
            processAndResult.AM3352_Version = am3352_version

            # ??????????????????
            HasRow = True
            command_seq_num_current = 1
            detail_table_head_row_num = coordinate.Detail_Table_Head_Row_Num # ????????????
            start_col_num = coordinate.Start_Col_Num # ????????????
            while HasRow:
                # ??????conmand??????
                current_row = detail_table_head_row_num + command_seq_num_current
                command_execute_times = processAndResult.Test_Circulate_Number  # ???????????????????????????????????????????????????????????????1
                command_coordinate = CommandCoordinate(command_seq_num_current, detail_table_head_row_num)
                command_seq_num = sheet.cell(command_coordinate.Command_Seq_Num_Location[0],
                                             command_coordinate.Command_Seq_Num_Location[1]).value  # ??????
                execution_phase = sheet.cell(command_coordinate.Execution_Phase_Location[0],
                                             command_coordinate.Execution_Phase_Location[1]).value  # ????????????
                test_method = sheet.cell(command_coordinate.Test_Method_Location[0],
                                         command_coordinate.Test_Method_Location[1]).value  # ????????????
                command_name = sheet.cell(command_coordinate.Command_Name_Location[0],
                                          command_coordinate.Command_Name_Location[1]).value  # ????????????/??????
                expect_value = sheet.cell(command_coordinate.Expect_Value_Location[0],
                                          command_coordinate.Expect_Value_Location[1]).value  # ?????????
                wait_time = sheet.cell(command_coordinate.Wait_Time_Location[0],
                                       command_coordinate.Wait_Time_Location[1]).value  # ??????????????????

                if execution_phase == "":
                    break
                if execution_phase == "??????":
                    command_execute_times = processAndResult.Test_Circulate_Number
                else:
                    command_execute_times = 0 # ???????????????????????????????????????
                processCommand = ProcessCommand(command_seq_num, execution_phase, test_method, command_name,
                                                expect_value, wait_time, command_execute_times, command_coordinate)
                logger.info(processCommand.__dict__)

                #  ????????? ??????????????????
                for i in range(processCommand.Execute_Times):
                    resultCoordinate = ResultCoordinate(current_row, i + 1, start_col_num)
                    processResult = ProcessResult(i + 1, resultCoordinate)
                    processCommand.Test_Results.append(processResult)
                processAndResult.Commands.append(processCommand)
                if execution_phase == "??????":
                    HasRow = False
                command_seq_num_current += 1
            return processAndResult
        else:
            logger.info("can not find sheet : " + coordinate.Sheet_Name)
            return None


if __name__ == '__main__':
    excel_template_file_name = "TCsetting_Template.xlsx"
    file_path = GetExcelTemplateFullPath(excel_template_file_name)
    excel_workbook = openpyxl.load_workbook(file_path)  # ??????????????????
    excelRead = ExcelRead(excel_workbook)
    processAndResul = excelRead.ReadProcessAndResultInfo()
    logger.info(processAndResul.__dict__)
    excel_workbook.close()
