import logging
import sys
import time
from openpyxl import Workbook
from ping3 import ping

sys.path.append("..")
sys.path.append("../..")
from log.logger import logger


class PlcConfigInfo:

    def __init__(self, ip_pc_rx65: str,
                 ip_pc_am3352: str,
                 ip_plc_rx65: str,
                 port_plc_rx65: int,
                 ip_plc_am3352: str,
                 ip_plc_am3352_check_mode: str,
                 port_plc_am3352_check_mode: int
                 ):
        self.Ip_pc_rx65 = ip_pc_rx65
        self.Ip_pc_am3352 = ip_pc_am3352
        self.Ip_plc_rx65 = ip_plc_rx65
        self.Port_plc_rx65 = port_plc_rx65
        self.Ip_plc_am3352 = ip_plc_am3352
        self.Ip_plc_am3352_check_mode = ip_plc_am3352_check_mode
        self.Port_plc_am3352_check_mode = port_plc_am3352_check_mode

    def GetConfigFromExcel(self, workbook: Workbook, sheet_name: str):
        sheets = workbook.sheetnames
        if sheets.__contains__(sheet_name):
            sheet = workbook[sheet_name]
            logger.info(sheet)
            self.Ip_pc_rx65 = sheet.cell(3, 3).value
            self.Ip_pc_am3352 = sheet.cell(4, 3).value
            self.Ip_plc_rx65 = sheet.cell(6, 3).value
            self.Port_plc_rx65 = sheet.cell(7, 3).value
        else:
            logger.info("配置文件中未找到名称为：{}的工作页".format(sheet_name))
            # logger.info("error异常")

    def CheckIpByPing(self, ip_to_check: str, check_times=10):
        host = ip_to_check
        while True:
            if check_times <= 0:
                return False
            logger.info("check: {}".format(host))
            res = ping(host)
            logger.info(res)
            if res is not None and res is not False:
                # logger.info('ping-{}成功，耗时{}s'.format(host, res))
                logger.info('check-{}成功，耗时{}s'.format(host, res))
                # break
                return True
            time.sleep(1)
            check_times = check_times - 1
        # logger.info("检测Ip：{  - 完成".format(host))

    def CheckIP_rx65(self):
        if self.Ip_plc_rx65 != "":
            return self.CheckIpByPing(self.Ip_plc_rx65 , check_times=20)
        else:
            return False

    def CheckIP_am3352_check_mode(self):
        if self.Ip_plc_am3352_check_mode != "":
            return self.CheckIpByPing(self.Ip_plc_am3352_check_mode, check_times=20)
        else:
            return False


if __name__ == '__main__':
    print("")
