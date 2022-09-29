import datetime
import sys

from openpyxl.worksheet.worksheet import Worksheet

sys.path.append("../..")
sys.path.append("../../..")
from models.cxe_click.CordinateCxeClick import CordinateCxeClick
from models.cxe_click.ClickCommandModel import ClickCommandModel
from models.cxe_click.SingleClickModel import SingleClickModel
from excel_tools.ExcelCommon import SetBodyCellStyle, SetHeaderCellStyle

from log.logger import logger


class CxeClickModel:

    def __init__(self, work_sheet: Worksheet, cordinate_cxe_click: CordinateCxeClick):
        self.WorkSheet = work_sheet  # Excel工作页
        self.Cordinate_Excel = cordinate_cxe_click  # Excel配置坐标

        self.ClickCmds: list[ClickCommandModel] = []  # 点击的命令集合
        self.Test_Cmd_Summary = 0  # 测试项目总数目
        self.Test_Cmd_Success = 0  # 测试项目成功数目
        self.Test_Cmd_Failed = 0  # 测试项目失败数目
        self.Test_Result = "OK"  # 测试结果
        self.TestCircleTime = None  # 测试循环次数

    # 写入excel
    def WriteCmdTestResultToExcel(self):
        # (1) 写如单次测试结果
        # (2) 写入单行汇总结果
        # (3) 写入全部汇总结果
        for ClickCmd in self.ClickCmds:
            for ClickModel in ClickCmd.ClickModels:
                ClickModel.WriteCurrentTestResultToExcel()
            ClickCmd.WriteCmdTestResultToExcel()

        # 计算并写入测试结果
        test_Cmd_Summary = 0
        test_Cmd_Failed = 0
        for each_test in self.ClickCmds:
            test_Cmd_Summary = test_Cmd_Summary + 1
            if each_test.CmdTestResult == "NG":
                test_Cmd_Failed = test_Cmd_Failed + 1
                self.Test_Result = "NG"

        # 计算并写入 测试项目成功数目
        self.Test_Cmd_Summary = test_Cmd_Summary
        self.Test_Cmd_Failed = test_Cmd_Failed
        self.Test_Cmd_Success = test_Cmd_Summary - test_Cmd_Failed

        # 写入excel，测试项目总数
        test_cmd_summary_cell = self.WorkSheet.cell(self.Cordinate_Excel.Test_Cmd_Summary_Location[0],
                                                    self.Cordinate_Excel.Test_Cmd_Summary_Location[1])
        test_cmd_summary_cell.value = self.Test_Cmd_Summary
        SetBodyCellStyle(test_cmd_summary_cell)

        # 写入excel，测试项目成功数
        test_cmd_success_cell = self.WorkSheet.cell(self.Cordinate_Excel.Test_Cmd_Success_Location[0],
                                                    self.Cordinate_Excel.Test_Cmd_Success_Location[1])
        test_cmd_success_cell.value = self.Test_Cmd_Success
        SetBodyCellStyle(test_cmd_success_cell)

        # 写入excel，测试项目失败数
        test_cmd_failed_cell = self.WorkSheet.cell(self.Cordinate_Excel.Test_Cmd_Failed_Location[0],
                                                   self.Cordinate_Excel.Test_Cmd_Failed_Location[1])
        test_cmd_failed_cell.value = self.Test_Cmd_Failed
        SetBodyCellStyle(test_cmd_failed_cell)

        # 写入excel，测试结果
        test_result_cell = self.WorkSheet.cell(self.Cordinate_Excel.Test_Result_Location[0],
                                               self.Cordinate_Excel.Test_Result_Location[1])
        test_result_cell.value = self.Test_Result
        SetBodyCellStyle(test_result_cell)

    # 初始化模型
    def InitModels(self):
        self.TestCircleTime = self.WorkSheet.cell(self.Cordinate_Excel.Test_Cycle_Time_Location[0],
                                                  self.Cordinate_Excel.Test_Cycle_Time_Location[1]).value
        # 读取excel中测试项目
        execute_cmd_seq = 1  # 项目序列号
        start_row = self.Cordinate_Excel.Test_Process_Header_Row + 1  # 明细数据起始行号
        current_row = start_row
        current_col = self.Cordinate_Excel.Test_Cmd_Col
        has_row = True
        while has_row:
            test_name = self.WorkSheet.cell(current_row, current_col).value
            if test_name is None or test_name == "":
                has_row = False
                break
            else:
                test_name = test_name.strip()
            # 初始化 ClickCommandModel
            clickCommandModel = ClickCommandModel(self.WorkSheet, self.Cordinate_Excel, test_name, execute_cmd_seq)
            for i in range(self.TestCircleTime):
                singleClickModel = SingleClickModel(self.WorkSheet, self.Cordinate_Excel, test_name, execute_cmd_seq, i+1)
                clickCommandModel.ClickModels.append(singleClickModel)
            self.ClickCmds.append(clickCommandModel)

            current_row = current_row + 1
            execute_cmd_seq = execute_cmd_seq + 1

    # 初始化Excel明细表头信息
    def InitTableHeader(self):
        wb = self.WorkSheet
        add_col_nums = self.TestCircleTime  # 测量循环次数， 添加的表头列数
        real_val_header_text_model = "实际值"
        test_result_text_model = "测试结果"
        start_col_num = self.Cordinate_Excel.Test_Process_Detail_Start_Col  # 起始列数 结果测量,此处默认5
        detail_table_head_row_num = self.Cordinate_Excel.Test_Process_Header_Row  # 表头行号
        for i in range(add_col_nums):
            real_val_col_num = start_col_num + i * 2
            test_result_col_num = start_col_num + i * 2 + 1
            real_val_text = real_val_header_text_model + (i + 1).__str__()
            test_result_text = test_result_text_model + (i + 1).__str__()
            real_val_cell = wb.cell(detail_table_head_row_num, real_val_col_num)
            test_result_cell = wb.cell(detail_table_head_row_num, test_result_col_num)
            real_val_cell.value = real_val_text
            test_result_cell.value = test_result_text
            SetHeaderCellStyle(real_val_cell)
            SetHeaderCellStyle(test_result_cell)


if __name__ == '__main__':
    logger.info("main")













